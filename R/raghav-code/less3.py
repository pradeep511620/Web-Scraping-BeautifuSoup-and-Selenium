import requests
import pandas as pd
from lxml import html
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import load_workbook
from openpyxl import Workbook
import numpy as np

dest_filename = r"D:\33.xlsx"

# to make sure the file exist
wb = Workbook()
wb.save(dest_filename)


def load_excel():
    wb = load_workbook(filename=dest_filename)
    ws = wb.active
    return wb, ws


def save_excel(all_df, ws):
    for r in dataframe_to_rows(all_df, index=False, header=False):
        ws.append(r)


# after one search page the file will be saved to excel.
def save_in_file(wb):
    wb.save(dest_filename)
def collect_metadata(t):
    product_name = t.xpath("//*[@id='product-wrapper']/div[1]/div/div/div[2]/div[1]/p/text()")
    price = t.xpath("//span[@itemprop='price']/text()")
    brand = t.xpath("//*[@id='product-wrapper']/div[1]/div/div/div[2]/h1/a/text()")
    item = t.xpath("//*[@id='product-wrapper']/div[1]/div/div/div[2]/div[2]/text()")

    mpn = t.xpath("//*[@id='product-wrapper']/div[1]/div/div/div[2]/h1/span/text()")
    image_url = t.xpath("//*[@id='product-wrapper']/div[1]/div/div/div[1]/figure/div/div[1]/div/div/img/@src")
    product_details1=t.xpath("//*[@id='details']/div/div[1]/ul/li[1]/text()")
    product_details2 = t.xpath("//*[@id='details']/div/div[1]/ul/li[2]/text()")
    product_details3 = t.xpath("//*[@id='details']/div/div[1]/ul/li[3]/text()")
    product_details4 = t.xpath("//*[@id='details']/div/div[1]/ul/li[4]/text()")
    product_details5 = t.xpath("//*[@id='details']/div/div[1]/ul/li[5]/text()")
    product_details6 = t.xpath("//*[@id='details']/div/div[1]/ul/li[6]/text()")
    aa = t.xpath("//table[contains(@class,'table-striped')]/tbody/tr/td[text()='Height']/./following-sibling::td/text()")
    bb = t.xpath("//table[@class='table table-striped table-striped--inverted table-stacked']//tr[1]/td[1]/text()")

    meta_data = [
        ("product_name", product_name[0] if product_name != [] else ""),
        ("price", price[0] if price != [] else ""),
        ("brand", brand[0] if brand != [] else ""),
        ("item", item[0] if item != [] else ""),
        ("mpn", mpn[0] if mpn != [] else ""),
        ("image_url", image_url[0] if image_url != [] else ""),
        ("Product Details", product_details1[0] if product_details1 != [] else ""),
        ("Product Details", product_details2[0] if product_details2 != [] else ""),
        ("Product Details", product_details3[0] if product_details3 != [] else ""),
        ("Product Details", product_details4[0] if product_details4 != [] else ""),
        ("Product Details", product_details5[0] if product_details5 != [] else ""),
        ("Product Details", product_details6[0] if product_details6 != [] else ""),
        (aa[0] if aa != [] else "", bb[0] if bb != [] else ""),
    ]

    return meta_data



def main():
   # th_xpath = "./thead/tr/th/text() | ./tbody/tr/th/text()"
   # td_xpath = "./thead/tr/td/text() | ./tbody/tr/td/text()"

    for page in range(0, 127):
        r = requests.get(
            f"https://www.applied.com/search?q=%3Arelevance%3Abrand%3ABrand-1288&page={page}&override=true&isLevelUp=false")  # https://www.mrosupply.com/brands/leeson/?page=1
        tree = html.fromstring(r.content)
        links = ["https://www.applied.com" + href for href in
                 tree.xpath(
                     "//div[@class='col-md-4 col-sm-3 product__image-wrapper']/a/@href")]  # //div[@class='product__image-wrapper']/a/@href
        wb, ws = load_excel()

        for link in links:
            req = requests.get(link)

            t = html.fromstring(req.text)

           # tables = t.xpath("//table[@class='description-table']")

            rows = []

            meta_data = collect_metadata(t)
            rows.extend(meta_data)


            all_df = pd.DataFrame(rows)
            all_df.insert(0, 'url', link)
            save_excel(all_df, ws)

        save_in_file(wb)
        print(f"search page - {page} scraped and saved, ")


if __name__ == "__main__":
    main()
