import requests
import pandas as pd
from lxml import html
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import load_workbook
from openpyxl import Workbook
import numpy as np

dest_filename = r"D:\hi.xlsx"

# to make sure the file exist
wb = Workbook()
wb.save(dest_filename)


def load_excel():
    wb = load_workbook(filename=dest_filename)
    ws = wb.active
    return wb, ws


def save_excel(all_df, ws):
    for r in dataframe_to_rows(all_df, index=False, header=False):
        ws.append(r)


# after one search page the file will be saved to excel.
def save_in_file(wb):
    wb.save(dest_filename)


def collect_metadata(t):
    product_name = t.xpath("//meta[@name='description']/@content")
    price = t.xpath("//p[@itemprop='price']/text()")
    brand = t.xpath("//*[@itemprop='brand']//span/text()")
    model = t.xpath("//*[@itemprop='mpn']/@content")
    condition = t.xpath("//li[contains(text(),'Condition: ')]/span/text()")
    sku = t.xpath("//*[@itemprop='sku']/text()")
    image_url = t.xpath("//*[@itemprop='image']/@src")


    meta_data = [
        ("product_name", product_name[0] if product_name != [] else ""),
        ("price", price[0] if price != [] else ""),
        ("brand", brand[0] if brand != [] else ""),
        ("model", model[0] if model != [] else ""),
        ("condition", condition[0] if condition != [] else ""),
        ("sku", sku[0] if sku != [] else ""),
        ("image_url", image_url[0] if image_url != [] else "")
    ]

    return meta_data


def main():
    th_xpath = "./thead/tr/th/text() | ./tbody/tr/th/text()"

    td_xpath = "./thead/tr/td/text() | ./tbody/tr/td/text()"

    for page in range(1, 113):
        r = requests.get(f"https://www.mrosupply.com/brands/marathon-electric/?page={page}")#https://www.mrosupply.com/brands/leeson/?page=1
        tree = html.fromstring(r.text)
        links = ["https://www.mrosupply.com" + href for href in
                 tree.xpath("//div[@class='m-catalogue-product-img']/a/@href")]
        wb, ws = load_excel()

        for link in links:
            req = requests.get(link)

            t = html.fromstring(req.text)
            tables = t.xpath("//table[@class='description-table']")

            rows = []

            meta_data = collect_metadata(t)
            rows.extend(meta_data)

            for table in tables:
                try:
                    if table.xpath("./thead/tr")[0].xpath("./td") == [] and table.xpath("./thead/tr")[0].xpath(
                            "./th") != []:
                        pass
                    else:
                        for a, b in zip(table.xpath(th_xpath), table.xpath(td_xpath)):
                            rows.append(tuple((a, b)))
                except IndexError:
                    for a, b in zip(table.xpath(th_xpath), table.xpath(td_xpath)):
                        rows.append(tuple((a, b)))
            all_df = pd.DataFrame(rows)
            all_df.insert(0, 'url', link)
            save_excel(all_df, ws)

        save_in_file(wb)
        print(f"search page - {page} scraped and saved, ")


if __name__ == "__main__":
    main()
