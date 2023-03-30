import requests
import pandas as pd
from lxml import html
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import load_workbook
from openpyxl import Workbook
import numpy as np

dest_filename = r"D:\33.xlsx"

# to make sure the file exist
wb = Workbook()
wb.save(dest_filename)


def load_excel():
    wb = load_workbook(filename=dest_filename)
    ws = wb.active
    return wb, ws


def save_excel(all_df, ws):
    for r in dataframe_to_rows(all_df, index=False, header=False):
        ws.append(r)


# after one search page the file will be saved to excel.
def save_in_file(wb):
    wb.save(dest_filename)


def collect_metadata(t):
    product_name = t.xpath("//meta[@name='description']/@content")
    price = t.xpath("//p[@itemprop='price']/text()")
    brand = t.xpath("//*[@itemprop='brand']//span/text()")
    model = t.xpath("//*[@itemprop='mpn']/@content")
    condition = t.xpath("//li[contains(text(),'Condition: ')]/span/text()")
    sku = t.xpath("//*[@itemprop='sku']/text()")
    image_url = t.xpath("//*[@itemprop='image']/@src")
    aa = t.xpath("//*[@id='additionalDescription']/div/div[2]/ul[2]/li[1]/text()")
    bb = t.xpath("//*[@id='additionalDescription']/div/div[2]/ul[2]/li[2]/text()")
    cc = t.xpath("//*[@id='additionalDescription']/div/div[2]/ul[2]/li[3]/text()")
    dd = t.xpath("//*[@id='additionalDescription']/div/div[2]/ul[2]/li[4]/text()")
    ee = t.xpath("//*[@id='additionalDescription']/div/div[2]/ul[2]/li[5]/text()")
    ff = t.xpath("//*[@id='additionalDescription']/div/div[2]/ul[2]/li[6]/text()")
    gg = t.xpath("//*[@id='additionalDescription']/div/div[2]/ul[2]/li[7]/text()")
    hh = t.xpath("//*[@id='additionalDescription']/div/div[2]/ul[2]/li[8]/text()")
    ii = t.xpath("//*[@id='additionalDescription']/div/div[2]/ul[2]/li[9]/text()")
    jj = t.xpath("//*[@id='additionalDescription']/div/div[2]/ul[2]/li[10]/text()")
    kk = t.xpath("//*[@id='additionalDescription']/div/div[2]/ul[2]/li[11]/text()")
    ll = t.xpath("//*[@id='additionalDescription']/div/div[2]/ul[2]/li[12]/text()")
    mm = t.xpath("//*[@id='additionalDescription']/div/div[2]/ul[2]/li[13]/text()")
    nn = t.xpath("//*[@id='additionalDescription']/div/div[2]/ul[2]/li[14]/text()")
    oo = t.xpath("//*[@id='additionalDescription']/div/div[2]/ul[2]/li[15]/text()")
    pp = t.xpath("//*[@id='additionalDescription']/div/div[2]/ul[2]/li[16]/text()")
    qq = t.xpath("//*[@id='additionalDescription']/div/div[2]/ul[2]/li[17]/text()")
    rr = t.xpath("//*[@id='additionalDescription']/div/div[2]/ul[2]/li[18]/text()")
    ss = t.xpath("//*[@id='additionalDescription']/div/div[2]/ul[2]/li[19]/text()")
    tt = t.xpath("//*[@id='additionalDescription']/div/div[2]/ul[2]/li[20]/text()")
    uu = t.xpath("//*[@id='additionalDescription']/div/div[2]/ul[2]/li[21]/text()")
    vv = t.xpath("//*[@id='additionalDescription']/div/div[2]/ul[2]/li[22]/text()")
    ww = t.xpath("//*[@id='additionalDescription']/div/div[2]/ul[2]/li[23]/text()")
    xx = t.xpath("//*[@id='additionalDescription']/div/div[2]/ul[2]/li[24]/text()")
    yy = t.xpath("//*[@id='additionalDescription']/div/div[2]/ul[2]/li[25]/text()")
    zz = t.xpath("//*[@id='additionalDescription']/div/div[2]/ul[2]/li[26]/text()")


    meta_data = [
        ("product_name", product_name[0] if product_name != [] else ""),
        ("price", price[0] if price != [] else ""),
        ("brand", brand[0] if brand != [] else ""),
        ("model", model[0] if model != [] else ""),
        ("condition", condition[0] if condition != [] else ""),
        ("sku", sku[0] if sku != [] else ""),
        ("image_url", image_url[0] if image_url != [] else ""),
        ("Attribute", aa[0] if aa != [] else ""),
        ("Attribute", bb[0] if bb != [] else ""),
        ("Attribute", cc[0] if cc != [] else ""),
        ("Attribute", dd[0] if dd != [] else ""),
        ("Attribute", ee[0] if ee != [] else ""),
        ("Attribute", ff[0] if ff != [] else ""),
        ("Attribute", gg[0] if gg != [] else ""),
        ("Attribute", hh[0] if hh != [] else ""),
        ("Attribute", ii[0] if ii != [] else ""),
        ("Attribute", jj[0] if jj != [] else ""),
        ("Attribute", kk[0] if kk != [] else ""),
        ("Attribute", ll[0] if ll != [] else ""),
        ("Attribute", mm[0] if mm != [] else ""),
        ("Attribute", nn[0] if nn != [] else ""),
        ("Attribute", oo[0] if oo != [] else ""),
        ("Attribute", pp[0] if pp != [] else ""),
        ("Attribute", qq[0] if qq != [] else ""),
        ("Attribute", rr[0] if rr != [] else ""),
        ("Attribute", ss[0] if ss != [] else ""),
        ("Attribute", tt[0] if tt != [] else ""),
        ("Attribute", uu[0] if uu != [] else ""),
        ("Attribute", vv[0] if vv != [] else ""),
        ("Attribute", ww[0] if ww != [] else ""),
        ("Attribute", xx[0] if xx != [] else ""),
        ("Attribute", yy[0] if yy != [] else ""),
        ("Attribute", zz[0] if zz != [] else ""),




    ]

    return meta_data


def main():
   # th_xpath = "./thead/tr/th/text() | ./tbody/tr/th/text()"
   # td_xpath = "./thead/tr/td/text() | ./tbody/tr/td/text()"

    for page in range(1, 127):
        r = requests.get(f"https://www.mrosupply.com/brands/marathon-electric/?page={page}")#https://www.mrosupply.com/brands/leeson/?page=1
        tree = html.fromstring(r.text)
        links = ["https://www.mrosupply.com" + href for href in
                 tree.xpath("//div[@class='m-catalogue-product-img']/a/@href")]
        wb, ws = load_excel()

        for link in links:
            req = requests.get(link)

            t = html.fromstring(req.text)
           # tables = t.xpath("//table[@class='description-table']")

            rows = []

            meta_data = collect_metadata(t)
            rows.extend(meta_data)


            all_df = pd.DataFrame(rows)
            all_df.insert(0, 'url', link)
            save_excel(all_df, ws)

        save_in_file(wb)
        print(f"search page - {page} scraped and saved, ")


if __name__ == "__main__":
    main()
